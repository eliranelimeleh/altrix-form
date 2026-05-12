from flask import Flask, request, jsonify, send_from_directory, send_file, session, redirect, url_for
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os, sqlite3, io, threading, webbrowser, smtplib, time
import secrets as _sec
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from zoneinfo import ZoneInfo

ISRAEL_TZ = ZoneInfo("Asia/Jerusalem")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or _sec.token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=bool(os.environ.get("FLY_APP_NAME")),
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE  = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "registrations.db"))
HEADERS  = ["שם מלא", "מס' טלפון", "אימייל", "סוג החשבון", "מס' חשבון מסחר", "תאריך ושעה"]


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    if os.path.dirname(DB_FILE):
        os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS registrations (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name     TEXT NOT NULL,
                phone         TEXT NOT NULL,
                email         TEXT NOT NULL,
                account_type  TEXT NOT NULL,
                trading_acct  TEXT NOT NULL,
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS client_updates (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name     TEXT NOT NULL,
                phone         TEXT NOT NULL,
                email         TEXT NOT NULL,
                action_type   TEXT NOT NULL,
                account_type  TEXT NOT NULL,
                new_account   TEXT NOT NULL,
                created_at    TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                created_date     TEXT NOT NULL,
                full_name        TEXT NOT NULL,
                phone            TEXT NOT NULL,
                trial_status     TEXT NOT NULL DEFAULT 'לפני ניסיון',
                lead_source      TEXT DEFAULT '',
                treatment_status TEXT NOT NULL DEFAULT 'חדש',
                notes            TEXT DEFAULT '',
                next_task        TEXT DEFAULT '',
                callback_time    TEXT DEFAULT '',
                is_closed        INTEGER NOT NULL DEFAULT 0,
                deal_amount      REAL DEFAULT NULL,
                payment_type     TEXT DEFAULT '',
                updated_at       TEXT DEFAULT '',
                agent            TEXT NOT NULL DEFAULT 'sales'
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sales_reps (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name  TEXT NOT NULL,
                phone      TEXT NOT NULL,
                start_date TEXT NOT NULL,
                email      TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        try:
            conn.execute("ALTER TABLE leads ADD COLUMN agent TEXT NOT NULL DEFAULT 'sales'")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE leads ADD COLUMN rep_id INTEGER DEFAULT NULL")
        except Exception:
            pass


def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "רישומים"
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor="1565C0")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = brd
    for i, w in enumerate([22, 18, 30, 18, 22, 24], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.row_dimensions[1].height = 28
    for rn, row in enumerate(rows, 2):
        alt = PatternFill("solid", fgColor="EBF3FB") if rn % 2 == 0 else None
        for i, v in enumerate([row["full_name"], row["phone"], row["email"],
                                row["account_type"], row["trading_acct"], row["created_at"]], 1):
            c = ws.cell(row=rn, column=i, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
            if alt: c.fill = alt
        ws.row_dimensions[rn].height = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def send_email_report(to_email, rows):
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    if not smtp_user or not smtp_pass:
        raise ValueError("SMTP credentials not configured")

    now   = datetime.now(ISRAEL_TZ)
    today = now.strftime("%d/%m/%Y")
    week_ago    = now - timedelta(days=7)
    month_start = now.replace(day=1, hour=0, minute=0, second=0)

    def parse(s):
        d,m,y,h,mn,sc = s.replace(" ","/").replace(":","/").split("/")
        return datetime(int(y),int(m),int(d),int(h),int(mn),int(sc), tzinfo=ISRAEL_TZ)

    today_count = sum(1 for r in rows if r["created_at"].startswith(today))
    week_count  = sum(1 for r in rows if parse(r["created_at"]) >= week_ago)
    month_count = sum(1 for r in rows if parse(r["created_at"]) >= month_start)

    type_counts = {}
    for r in rows:
        type_counts[r["account_type"]] = type_counts.get(r["account_type"], 0) + 1
    type_rows = "".join(
        f"<tr><td style='padding:8px 14px;border-bottom:1px solid #e5e7eb'>{t}</td>"
        f"<td style='padding:8px 14px;border-bottom:1px solid #e5e7eb;text-align:center'>{c}</td>"
        f"<td style='padding:8px 14px;border-bottom:1px solid #e5e7eb;text-align:center'>{round(c/len(rows)*100) if rows else 0}%</td></tr>"
        for t, c in sorted(type_counts.items(), key=lambda x: -x[1])
    )

    recent_rows = "".join(
        f"<tr><td style='padding:8px 14px;border-bottom:1px solid #e5e7eb'>{r['full_name']}</td>"
        f"<td style='padding:8px 14px;border-bottom:1px solid #e5e7eb'>{r['phone']}</td>"
        f"<td style='padding:8px 14px;border-bottom:1px solid #e5e7eb'>{r['account_type']}</td>"
        f"<td style='padding:8px 14px;border-bottom:1px solid #e5e7eb'>{r['created_at']}</td></tr>"
        for r in list(reversed(rows))[:10]
    )

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:620px;margin:0 auto;direction:rtl">
      <div style="background:linear-gradient(135deg,#0d1b3e,#1565c0);padding:32px;border-radius:12px 12px 0 0;text-align:center">
        <h1 style="color:#fff;margin:0;font-size:1.8rem;letter-spacing:3px">ALTRIX</h1>
        <p style="color:rgba(255,255,255,.7);margin:8px 0 0;font-size:.9rem">סיכום רישומים — {today}</p>
      </div>
      <div style="background:#f9fafb;padding:28px;border-radius:0 0 12px 12px;border:1px solid #e5e7eb">

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-bottom:28px">
          {''.join(f'<div style="background:#fff;border-radius:10px;padding:16px;text-align:center;border:1px solid #e5e7eb"><div style="font-size:.75rem;color:#6b7280;font-weight:600;margin-bottom:6px">{lbl}</div><div style="font-size:1.8rem;font-weight:800;color:{col}">{val}</div></div>'
            for lbl,val,col in [("סה\"כ",len(rows),"#1565c0"),("היום",today_count,"#2e7d32"),("השבוע",week_count,"#e65100"),("החודש",month_count,"#6a1b9a")])}
        </div>

        <h3 style="font-size:.95rem;color:#111827;margin:0 0 12px">התפלגות סוגי חשבונות</h3>
        <table style="width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;border:1px solid #e5e7eb;margin-bottom:24px">
          <thead><tr style="background:#1565c0">
            <th style="padding:10px 14px;color:#fff;text-align:right;font-size:.82rem">סוג חשבון</th>
            <th style="padding:10px 14px;color:#fff;text-align:center;font-size:.82rem">כמות</th>
            <th style="padding:10px 14px;color:#fff;text-align:center;font-size:.82rem">אחוז</th>
          </tr></thead>
          <tbody>{type_rows}</tbody>
        </table>

        <h3 style="font-size:.95rem;color:#111827;margin:0 0 12px">10 רישומים אחרונים</h3>
        <table style="width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;border:1px solid #e5e7eb;margin-bottom:24px">
          <thead><tr style="background:#1565c0">
            <th style="padding:10px 14px;color:#fff;text-align:right;font-size:.82rem">שם</th>
            <th style="padding:10px 14px;color:#fff;text-align:right;font-size:.82rem">טלפון</th>
            <th style="padding:10px 14px;color:#fff;text-align:right;font-size:.82rem">סוג חשבון</th>
            <th style="padding:10px 14px;color:#fff;text-align:right;font-size:.82rem">תאריך</th>
          </tr></thead>
          <tbody>{recent_rows}</tbody>
        </table>

        <p style="font-size:.8rem;color:#9ca3af;text-align:center;margin:0">
          נשלח אוטומטית מ-ALTRIX Dashboard • {now.strftime('%d/%m/%Y %H:%M')}
        </p>
      </div>
    </div>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"Altrix — סיכום רישומים {today}"
    msg["From"]    = smtp_user
    msg["To"]      = to_email
    msg.attach(MIMEText(html, "html", "utf-8"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(smtp_user, smtp_pass)
        s.sendmail(smtp_user, to_email, msg.as_bytes())


def is_authed():
    key = os.environ.get("DOWNLOAD_KEY", "Ariel123")
    return session.get("authed") or request.args.get("key") == key

def require_auth():
    if not is_authed():
        if request.path.startswith("/api/"):
            return jsonify({"error": "unauthorized"}), 401
        return redirect("/login")
    return None

# ── Manager security ──────────────────────────────────────────────────────────

# MANAGER_KEY must be set as a Fly secret; if missing, login is disabled
MANAGER_KEY: str = os.environ.get("MANAGER_KEY") or _sec.token_hex(32)

# Rate limiting: max 5 attempts per IP per 15 minutes
_mgr_attempts: dict = defaultdict(list)
_MGR_MAX = 5
_MGR_LOCKOUT = 900  # seconds


def _client_ip() -> str:
    xff = request.headers.get("X-Forwarded-For", "")
    return xff.split(",")[0].strip() or request.remote_addr or "unknown"


def _mgr_rate_check(ip: str):
    now = time.time()
    lst = _mgr_attempts[ip]
    lst[:] = [t for t in lst if now - t < _MGR_LOCKOUT]
    if len(lst) >= _MGR_MAX:
        wait = int(_MGR_LOCKOUT - (now - lst[0]))
        return False, wait
    return True, 0


def _mgr_record_fail(ip: str):
    _mgr_attempts[ip].append(time.time())


def _mgr_clear(ip: str):
    _mgr_attempts.pop(ip, None)


def is_manager_authed() -> bool:
    # Session uses obscured key names to prevent obvious forgery attempts
    if not session.get("_mgt"):
        return False
    token = session.get("_mgk", "")
    if not token or len(token) < 32:
        return False
    # IP binding: session cookie only valid from the IP that logged in
    bound_ip = session.get("_mgi")
    if bound_ip and bound_ip != _client_ip():
        app.logger.warning(f"Manager session IP mismatch: bound={bound_ip} current={_client_ip()}")
        return False
    return True


def require_manager_auth():
    if not is_manager_authed():
        session.clear()  # wipe any tampered/partial session state
        if request.path.startswith("/api/"):
            return jsonify({"error": "unauthorized"}), 401
        return redirect("/manager-login")
    return None


@app.after_request
def security_headers(response):
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), camera=(), microphone=()"
    if request.path.startswith("/manager"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
        response.headers["Pragma"] = "no-cache"
    return response


UPDATABLE_FIELDS = {
    "created_date", "full_name", "phone", "trial_status", "lead_source",
    "treatment_status", "notes", "next_task", "callback_time",
    "is_closed", "deal_amount", "payment_type"
}
MANAGER_UPDATABLE_FIELDS = UPDATABLE_FIELDS | {"rep_id"}


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/update-form")
def update_form():
    return send_from_directory(BASE_DIR, "update.html")


@app.route("/update", methods=["POST"])
def update():
    data = request.get_json(force=True)
    for f in ["fullName", "phone", "email", "actionType", "accountType", "newAccount"]:
        if not str(data.get(f, "")).strip():
            return jsonify({"error": f"missing: {f}"}), 400
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO client_updates (full_name,phone,email,action_type,account_type,new_account,created_at) VALUES (?,?,?,?,?,?,?)",
            (data["fullName"], data["phone"], data["email"],
             data["actionType"], data["accountType"], data["newAccount"], ts)
        )
    # WhatsApp notification for updates
    wa_phone  = os.environ.get("WA_PHONE", "")
    wa_apikey = os.environ.get("WA_APIKEY", "")
    if wa_phone and wa_apikey:
        try:
            msg = f"Altrix - בקשת עדכון חדשה!\n{data['actionType']}\nשם: {data['fullName']}\nטל: {data['phone']}\nסוג: {data['accountType']}\nחשבון: {data['newAccount']}"
            import urllib.request, urllib.parse
            url = f"https://api.callmebot.com/whatsapp.php?phone={wa_phone}&text={urllib.parse.quote(msg)}&apikey={wa_apikey}"
            urllib.request.urlopen(url, timeout=5)
        except:
            pass
    return jsonify({"ok": True}), 200


@app.route("/api/updates")
def api_updates():
    denied = require_auth()
    if denied: return denied
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM client_updates ORDER BY id").fetchall()
    return jsonify({"rows": [dict(r) for r in rows]})


@app.route("/dashboard")
def dashboard():
    denied = require_auth()
    if denied: return denied
    return send_from_directory(BASE_DIR, "dashboard.html")


@app.route("/api/stats")
def api_stats():
    denied = require_auth()
    if denied: return denied
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM registrations ORDER BY id").fetchall()
    return jsonify({"rows": [dict(r) for r in rows]})


@app.route("/submit", methods=["POST"])
def submit():
    data = request.get_json(force=True)
    for f in ["fullName", "phone", "email", "accountType", "tradingAccount"]:
        if not str(data.get(f, "")).strip():
            return jsonify({"error": f"missing: {f}"}), 400
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO registrations (full_name,phone,email,account_type,trading_acct,created_at) VALUES (?,?,?,?,?,?)",
            (data["fullName"], data["phone"], data["email"],
             data["accountType"], data["tradingAccount"], ts)
        )
    # WhatsApp notification for new registration
    wa_phone  = os.environ.get("WA_PHONE", "")
    wa_apikey = os.environ.get("WA_APIKEY", "")
    if wa_phone and wa_apikey:
        try:
            msg = f"Altrix - רישום חדש!\nשם: {data['fullName']}\nטל: {data['phone']}\nסוג: {data['accountType']}\nחשבון: {data['tradingAccount']}"
            import urllib.request, urllib.parse
            url = f"https://api.callmebot.com/whatsapp.php?phone={wa_phone}&text={urllib.parse.quote(msg)}&apikey={wa_apikey}"
            urllib.request.urlopen(url, timeout=5)
        except:
            pass
    return jsonify({"ok": True}), 200


@app.route("/download")
def download():
    denied = require_auth()
    if denied: return denied
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM registrations ORDER BY id").fetchall()
    buf = build_excel(rows)
    today = datetime.now(ISRAEL_TZ).strftime("%d-%m-%Y")
    return send_file(buf, as_attachment=True,
                     download_name=f"Altrix_{today}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/send-report", methods=["POST"])
def send_report():
    denied = require_auth()
    if denied: return denied
    data = request.get_json(force=True)
    to_email = data.get("email", "").strip()
    if not to_email:
        return jsonify({"error": "missing email"}), 400
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM registrations ORDER BY id").fetchall()
    try:
        send_email_report(to_email, rows)
        return jsonify({"ok": True}), 200
    except ValueError as e:
        return jsonify({"error": "המייל לא מוגדר בשרת — ראה הוראות הגדרה"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        pwd = request.form.get("password", "")
        if pwd == os.environ.get("DOWNLOAD_KEY", "altrix2024"):
            session["authed"] = True
            return redirect("/sales")
        return redirect("/login?err=1")
    return send_from_directory(BASE_DIR, "login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.route("/sales")
def sales():
    # Persist session when accessing with key URL param so API calls work
    key = os.environ.get("DOWNLOAD_KEY", "Ariel123")
    if request.args.get("key") == key:
        session["authed"] = True
    denied = require_auth()
    if denied: return denied
    return send_from_directory(BASE_DIR, "sales.html")


@app.route("/api/leads")
def api_leads():
    denied = require_auth()
    if denied: return denied
    closed_filter = request.args.get("closed")
    with get_db() as conn:
        if closed_filter is not None:
            rows = conn.execute(
                "SELECT * FROM leads WHERE agent='sales' AND is_closed=? ORDER BY id DESC",
                (int(closed_filter),)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM leads WHERE agent='sales' ORDER BY id DESC").fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["is_closed"] = bool(d["is_closed"])
        result.append(d)
    return jsonify({"leads": result})


@app.route("/api/leads", methods=["POST"])
def create_lead():
    denied = require_auth()
    if denied: return denied
    data = request.get_json(force=True)
    for f in ["full_name", "phone"]:
        if not str(data.get(f, "")).strip():
            return jsonify({"error": f"missing: {f}"}), 400
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    today_date = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y")
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO leads
               (created_date, full_name, phone, trial_status, lead_source,
                treatment_status, notes, next_task, callback_time,
                is_closed, deal_amount, payment_type, updated_at, agent)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                data.get("created_date", today_date),
                data["full_name"].strip(),
                data["phone"].strip(),
                data.get("trial_status", "לפני ניסיון"),
                data.get("lead_source", ""),
                data.get("treatment_status", "חדש"),
                data.get("notes", ""),
                data.get("next_task", ""),
                data.get("callback_time", ""),
                1 if data.get("is_closed") else 0,
                data.get("deal_amount"),
                data.get("payment_type", ""),
                ts,
                'sales'
            )
        )
        new_id = cur.lastrowid
    return jsonify({"ok": True, "id": new_id}), 201


@app.route("/api/leads/<int:lead_id>", methods=["PUT"])
def update_lead(lead_id):
    denied = require_auth()
    if denied: return denied
    data = request.get_json(force=True)
    fields = {k: v for k, v in data.items() if k in UPDATABLE_FIELDS}
    if not fields:
        return jsonify({"error": "no fields to update"}), 400
    if "is_closed" in fields:
        fields["is_closed"] = 1 if fields["is_closed"] else 0
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    fields["updated_at"] = ts
    set_clause = ", ".join(f"{k}=?" for k in fields)
    values = list(fields.values()) + [lead_id]
    with get_db() as conn:
        cur = conn.execute(f"UPDATE leads SET {set_clause} WHERE id=? AND agent='sales'", values)
        if cur.rowcount == 0:
            return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})


@app.route("/api/leads/<int:lead_id>", methods=["DELETE"])
def delete_lead(lead_id):
    denied = require_auth()
    if denied: return denied
    with get_db() as conn:
        cur = conn.execute("DELETE FROM leads WHERE id=? AND agent='sales'", (lead_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})


# ── Manager routes ────────────────────────────────────────────────────────────

@app.route("/manager-login", methods=["GET", "POST"])
def manager_login():
    if request.method == "POST":
        ip = _client_ip()
        allowed, wait = _mgr_rate_check(ip)
        if not allowed:
            m, s = divmod(wait, 60)
            app.logger.warning(f"Manager login rate-limited: {ip}")
            return redirect(f"/manager-login?err=locked&wait={m}:{s:02d}")

        pwd = request.form.get("password", "")
        # Constant-time comparison prevents timing-based password discovery
        if _sec.compare_digest(pwd.encode("utf-8"), MANAGER_KEY.encode("utf-8")):
            _mgr_clear(ip)
            session.clear()  # evict any existing session (e.g., stale sales auth)
            session["_mgt"] = True               # manager authenticated flag
            session["_mgk"] = _sec.token_hex(32) # per-session cryptographic token
            session["_mgi"] = ip                 # IP binding
            session.permanent = True
            app.logger.info(f"Manager login success: {ip}")
            return redirect("/manager")

        _mgr_record_fail(ip)
        remaining = _MGR_MAX - len(_mgr_attempts[ip])
        app.logger.warning(f"Manager login failed: {ip} ({len(_mgr_attempts[ip])}/{_MGR_MAX})")
        return redirect(f"/manager-login?err=1&left={max(remaining,0)}")

    return send_from_directory(BASE_DIR, "manager-login.html")


@app.route("/manager-logout")
def manager_logout():
    session.clear()
    return redirect("/manager-login")


@app.route("/manager")
def manager_dashboard():
    denied = require_manager_auth()
    if denied: return denied
    return send_from_directory(BASE_DIR, "manager.html")


@app.route("/api/manager/leads")
def api_manager_leads():
    denied = require_manager_auth()
    if denied: return denied
    closed_filter = request.args.get("closed")
    rep_filter = request.args.get("rep_id")
    with get_db() as conn:
        conditions = ["agent='manager'"]
        params = []
        if closed_filter is not None:
            conditions.append("is_closed=?")
            params.append(int(closed_filter))
        if rep_filter is not None:
            if rep_filter == "unassigned":
                conditions.append("rep_id IS NULL")
            else:
                conditions.append("rep_id=?")
                params.append(int(rep_filter))
        where = " AND ".join(conditions)
        rows = conn.execute(f"SELECT * FROM leads WHERE {where} ORDER BY id DESC", params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["is_closed"] = bool(d["is_closed"])
        result.append(d)
    return jsonify({"leads": result})


@app.route("/api/manager/leads", methods=["POST"])
def create_manager_lead():
    denied = require_manager_auth()
    if denied: return denied
    data = request.get_json(force=True)
    for f in ["full_name", "phone"]:
        if not str(data.get(f, "")).strip():
            return jsonify({"error": f"missing: {f}"}), 400
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    today_date = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y")
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO leads
               (created_date, full_name, phone, trial_status, lead_source,
                treatment_status, notes, next_task, callback_time,
                is_closed, deal_amount, payment_type, updated_at, agent)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                data.get("created_date", today_date),
                data["full_name"].strip(),
                data["phone"].strip(),
                data.get("trial_status", "לפני ניסיון"),
                data.get("lead_source", ""),
                data.get("treatment_status", "חדש"),
                data.get("notes", ""),
                data.get("next_task", ""),
                data.get("callback_time", ""),
                1 if data.get("is_closed") else 0,
                data.get("deal_amount"),
                data.get("payment_type", ""),
                ts,
                'manager'
            )
        )
        new_id = cur.lastrowid
    return jsonify({"ok": True, "id": new_id}), 201


@app.route("/api/manager/leads/<int:lead_id>", methods=["PUT"])
def update_manager_lead(lead_id):
    denied = require_manager_auth()
    if denied: return denied
    data = request.get_json(force=True)
    fields = {k: v for k, v in data.items() if k in MANAGER_UPDATABLE_FIELDS}
    if not fields:
        return jsonify({"error": "no fields to update"}), 400
    if "is_closed" in fields:
        fields["is_closed"] = 1 if fields["is_closed"] else 0
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    fields["updated_at"] = ts
    set_clause = ", ".join(f"{k}=?" for k in fields)
    values = list(fields.values()) + [lead_id]
    with get_db() as conn:
        cur = conn.execute(f"UPDATE leads SET {set_clause} WHERE id=? AND agent='manager'", values)
        if cur.rowcount == 0:
            return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})


@app.route("/api/manager/leads/<int:lead_id>", methods=["DELETE"])
def delete_manager_lead(lead_id):
    denied = require_manager_auth()
    if denied: return denied
    with get_db() as conn:
        cur = conn.execute("DELETE FROM leads WHERE id=? AND agent='manager'", (lead_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})


# ── Sales Reps API ───────────────────────────────────────────────────────────

@app.route("/api/manager/reps")
def api_manager_reps():
    denied = require_manager_auth()
    if denied: return denied
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM sales_reps ORDER BY full_name").fetchall()
    return jsonify({"reps": [dict(r) for r in rows]})


@app.route("/api/manager/reps", methods=["POST"])
def create_manager_rep():
    denied = require_manager_auth()
    if denied: return denied
    data = request.get_json(force=True)
    for f in ["full_name", "phone"]:
        if not str(data.get(f, "")).strip():
            return jsonify({"error": f"missing: {f}"}), 400
    ts = datetime.now(ISRAEL_TZ).strftime("%d/%m/%Y %H:%M:%S")
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO sales_reps (full_name, phone, start_date, email, created_at) VALUES (?,?,?,?,?)",
            (data["full_name"].strip(), data["phone"].strip(),
             data.get("start_date", ""), data.get("email", ""), ts)
        )
        new_id = cur.lastrowid
    return jsonify({"ok": True, "id": new_id}), 201


@app.route("/api/manager/reps/<int:rep_id>", methods=["DELETE"])
def delete_manager_rep(rep_id):
    denied = require_manager_auth()
    if denied: return denied
    with get_db() as conn:
        cur = conn.execute("DELETE FROM sales_reps WHERE id=?", (rep_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "not found"}), 404
        # Un-assign any leads that were assigned to this rep
        conn.execute("UPDATE leads SET rep_id=NULL WHERE rep_id=?", (rep_id,))
    return jsonify({"ok": True})


@app.route("/api/manager/stats")
def api_manager_stats():
    """Per-rep performance stats within an optional date range."""
    denied = require_manager_auth()
    if denied: return denied
    from_date = request.args.get("from_date", "")  # dd/mm/yyyy
    to_date   = request.args.get("to_date", "")
    with get_db() as conn:
        reps = conn.execute("SELECT * FROM sales_reps ORDER BY full_name").fetchall()
        leads = conn.execute("SELECT * FROM leads WHERE agent='manager'").fetchall()
    leads = [dict(l) for l in leads]

    def in_range(lead):
        d = lead.get("created_date", "")
        if not d: return True
        try:
            dd, mm, yy = d.split("/")
            ld = f"{yy}-{mm}-{dd}"
        except Exception:
            return True
        if from_date:
            try:
                fd, fm, fy = from_date.split("/")
                if ld < f"{fy}-{fm}-{fd}": return False
            except Exception:
                pass
        if to_date:
            try:
                td, tm, ty = to_date.split("/")
                if ld > f"{ty}-{tm}-{td}": return False
            except Exception:
                pass
        return True

    filtered = [l for l in leads if in_range(l)]
    stats = []
    for rep in reps:
        rep_leads = [l for l in filtered if l.get("rep_id") == rep["id"]]
        closed = [l for l in rep_leads if l.get("is_closed")]
        revenue = sum(l.get("deal_amount") or 0 for l in closed)
        stats.append({
            "id": rep["id"],
            "full_name": rep["full_name"],
            "email": rep["email"],
            "phone": rep["phone"],
            "start_date": rep["start_date"],
            "total_leads": len(rep_leads),
            "closed": len(closed),
            "revenue": revenue,
            "conversion": round(len(closed) / len(rep_leads) * 100, 1) if rep_leads else 0
        })
    unassigned = [l for l in filtered if l.get("rep_id") is None]
    closed_u = [l for l in unassigned if l.get("is_closed")]
    stats.append({
        "id": None,
        "full_name": "לא משויך",
        "email": "",
        "phone": "",
        "start_date": "",
        "total_leads": len(unassigned),
        "closed": len(closed_u),
        "revenue": sum(l.get("deal_amount") or 0 for l in closed_u),
        "conversion": round(len(closed_u) / len(unassigned) * 100, 1) if unassigned else 0
    })
    return jsonify({"stats": stats})


# ── Startup ───────────────────────────────────────────────────────────────────

def open_browser():
    webbrowser.open("http://localhost:5000")


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    if port == 5000:
        threading.Timer(1.2, open_browser).start()
    print(f"\nServer: http://localhost:{port}")
    print("Press Ctrl+C to stop\n")
    app.run(host="0.0.0.0", port=port)
